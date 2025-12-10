import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Create a Tkinter root window (this will be hidden)
root = tk.Tk()
root.withdraw()  # Hide the root window as we only need the file dialog

# Open the file selection dialog
file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=(("Excel files", "*.xlsx;*.xlsm"), ("All files", "*.*")))

# Check if the user selected a file
if file_path:
    # Load your Excel file
    workbook = load_workbook(file_path)
    sheet = workbook.active  # or use sheet = workbook['SheetName'] if you have multiple sheets

    # Find the column with the header "LAN IP"
    lan_ip_column = None
    for col in sheet.iter_cols(min_row=1, max_row=1):  # We are only checking the first row (headers)
        for cell in col:
            if cell.value == 'LAN IP':
                lan_ip_column = col[0].column  # This gives us the column number
                break
        if lan_ip_column:
            break

    if lan_ip_column is None:
        print("Column 'LAN IP' not found.")
    else:
        # Initialize a list to hold the values of colored cells
        colored_cells_values = []

        # Loop through the rows in the "LAN IP" column and check for any colored cells
        for row in sheet.iter_rows(min_col=lan_ip_column, max_col=lan_ip_column, min_row=2):  # Start from row 2 to skip headers
            for cell in row:
                # Print the actual color of the cell for debugging
                fill_color = cell.fill.start_color.index
                #print(f"Cell {cell.row} - Color: {fill_color}")  # Debug: print the color index of each cell

                # Check if the cell has a fill color (yellowish or any other desired color)
                if fill_color == 'FFFFC000':  # Example: check for specific color (yellowish)
                    colored_cells_values.append(cell.value)

        # Ping and report logic follows (same as before)
        import subprocess
        import re

        # Function to ping an IP address with a timeout
        def ping_ip(ip):
            try:
                print(f"Pinging {ip}...")  # Debugging step
                # Run the ping command, "-n 2" means 2 packets are sent
                result = subprocess.run(['ping', ip, '-n', '2'], stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=10, universal_newlines=True)

                # Print the output to see what's returned
                print("stdout:", result.stdout)  # Show standard output
                print("stderr:", result.stderr)  # Show standard error

                # Check if the ping command was successful by looking for 'Reply from'
                if ": bytes=32 time=" in result.stdout:
                    print(f"Ping to {ip} was successful.")
                    return True
                else:
                    print(f"Ping to {ip} failed: {result.stderr}")
                    return False

            except subprocess.TimeoutExpired:
                print(f"Ping to {ip} timed out.")  # Debugging step
                return False
            except Exception as e:
                print(f"Error pinging {ip}: {e}")  # Debugging step
                return False

        # Function to increment the last octet of an IP address
        def increment_ip(ip):
            # Match the IP address pattern (IPv4 format)
            match = re.match(r'(\d+\.\d+\.\d+)\.(\d+)', ip)
            if match:
                base_ip = match.group(1)
                last_octet = int(match.group(2))
                # Prevent incrementing the last octet beyond 255
                if last_octet < 255:
                    # Increment the last octet by 1
                    new_ip = f"{base_ip}.{last_octet + 1}"
                    return new_ip
            return None

        # Example list of IP addresses
        ip_addresses = colored_cells_values
        # Open a text file to write the report
        with open('ping_report.txt', 'w') as file:
            for ip in ip_addresses:
                # Ping the original IP address
                if ping_ip(ip):
                    file.write(f"{ip} is reachable.\n")
                    # Now increment the last octet and ping the new address
                    new_ip = increment_ip(ip)
                    if new_ip and ping_ip(new_ip):
                        file.write(f"{new_ip} is reachable.\n")
                    else:
                        file.write(f"{new_ip} is not reachable.\n")
                        file.write("_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_\n")
                        file.write("-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-\n")
                    
                else:
                    file.write(f"{ip} is not reachable.\n")
                    file.write("_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_\n")
                    file.write("-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-\n")
        print("Ping report has been written to 'ping_report.txt'.")
else:
    print("No file selected.")